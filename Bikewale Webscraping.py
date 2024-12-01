from selenium import webdriver
from selenium.webdriver.common.by import By
import html
import time
import requests
from bs4 import BeautifulSoup
from threading import Thread,BoundedSemaphore
import os
import shutil
import traceback
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException,StaleElementReferenceException,TimeoutException,ElementClickInterceptedException
import sys
import trace
import re
import pprint
from openpyxl import load_workbook
import xlsxwriter
import openpyxl
from openpyxl.styles import Font
import win32com.client as win32
import pandas as pd

Brands = ['kawasaki']#,'honda','hero','bajaj','triumph','bmw','tvs','ducati','yamaha','ktm','royalenfield','suzuki','harleydavidson','benelli','jawa','husqvarna','yezdi']

def threadtry(brand):
    limiter.acquire()
    class driverclass:
        def __init__(self) :
            
            self.listofbikes = '' 
            self.listofbikeback = ''
            self.lisofbikename = ''
            self.driver = ''
            self.ulparentclass = ''
            self.Places = ''
            self.row1 = 1
            self.lastbike = ''
            self.lastplace = ''
            self.url1 = ''
            self.reruncheck = ''
            self.workbook = ''
            self.sheet2 = ''
        def create_driver(self):
                    
            preferences = {"download.default_directory": r'C:\Users\abhishekkd\Downloads',  "download.directory_upgrade": True,
                        "download.prompt_for_download": False }

            options = webdriver.ChromeOptions()
            options.add_experimental_option("prefs", preferences )
            self.driver = webdriver.Chrome(options=options , executable_path= r"C:\Users\USER\Downloads\chromedriver-win32\chromedriver-win32\chromedriver.exe")
            return self.driver
        def mainfunc(self , driver,brand):
            if self.reruncheck == '' or self.reruncheck != 0:
                
                self.workbook = load_workbook(filename=r"C:\Users\USER\Documents\Competitor Price.xlsx")
                sheet1 = self.workbook['Sheet1']
                self.sheet2 = self.workbook[brand]
                driver.get("https://www.bikewale.com/"+brand+"-bikes/")
                
                wait = WebDriverWait(driver, 20 , ignored_exceptions=(NoSuchElementException,StaleElementReferenceException,ElementClickInterceptedException))
                # driver.maximize_window()
                wait.until(EC.url_contains(brand))
                driver.maximize_window()

                self.url1 = driver.current_url
                print(self.url1)

                source1 = requests.get(self.url1 )#,verify= r'c:\Users\abhishekkd\Downloads\cert\cacertcopy - Copy.pem').text
                # print(source1.text)
               
                soup = BeautifulSoup(source1.text, 'lxml')
                
                ratingicon = soup.find('svg' , attrs={"aria-label" : "rating icon"})
                
                print(type(ratingicon))
                ul = ratingicon.find_parent('ul')
                ulparent = ul.find_parent().find_parent()
                for i in ulparent.attrs['class']:
                    self.ulparentclass = i
                

                self.listofbikes = ul.find_all('li')
                self.listofbikename = []
                for i in self.listofbikes:
                    self.listofbikename.append(i.find('a')['title'])
                    for kl in i.descendants:
                        if kl.string == 'Estimated Price':
                            # print('yes')
                            self.listofbikename.remove(i.find('a')['title'])
                            self.listofbikes.remove(i)
                            break

                print(self.listofbikename)
                print(len(self.listofbikes))
                self.Places = []
                for cells in sheet1['A']:
                    self.Places.append(cells.value)
                
                self.listofbikeback = 1
                self.row1 = 1
                self.reruncheck = 1
                self.lastbike = self.listofbikes[0].find('a')['title']
                self.lastplace = self.Places[0]

            for j,bike in enumerate(self.listofbikes):
                if self.lastbike == bike.find('a')['title']:
                    bikeorder = j
                    print(j)
                    break
            for j,placename in enumerate(self.Places):
                if self.lastplace == self.Places[-1]:
                    placeorder = 0
                    break
                if placename == self.lastplace:
                    placeorder = j   
                    print(j,placename)
                    break

            
            for i in self.listofbikes[bikeorder:]:
                print(brand,' listofbikeback value is ',self.listofbikeback, ' ',self.reruncheck)
                print(brand,' length of bikelist ',len(self.listofbikes[bikeorder:]))
                if self.listofbikeback != 1 or self.reruncheck == 0:
                    wait = WebDriverWait(driver, 20 , ignored_exceptions=(NoSuchElementException,StaleElementReferenceException,ElementClickInterceptedException))

                    driver.get(self.url1)
                    
                    time.sleep(2)
                    wait.until(EC.url_contains(brand))
                    
                    url = driver.current_url
                    source1 = requests.get(url )#,verify= r'c:\Users\abhishekkd\Downloads\cert\cacertcopy - Copy.pem').text
                    print(i.find('a')['title'])
                    time.sleep(2)
                    driver.maximize_window()
                ulselenium =  wait.until(EC.presence_of_element_located((By.XPATH , '//div[@class = "'+self.ulparentclass+'" ]')))   

                liselenium = ulselenium.find_element(By.XPATH, './div[4]/ul/li['+str(self.listofbikeback)+']')
    

                print(i.find('a')['title'])

                try:
                    bike = liselenium.find_element(By.XPATH, './/button')
                except :
                    if liselenium.find_elements_by_tag_name('button'):
                        bike = liselenium.find_element(By.XPATH, './/button')    
                    else:
                        print(i.find('a')['title'],' bike not launched')
                        self.listofbikeback = self.listofbikeback + 1
                        continue
                try:
                    bike.click()
                except:
                    driver.execute_script("arguments[0].scrollIntoView();", bike)
                    driver.execute_script("arguments[0].click();", bike) 
                    
                if self.listofbikeback == 1 or self.reruncheck == 0:
                    detectlocation = wait.until(EC.presence_of_element_located((By.XPATH , '//p[text() = "Detect my location" ]')))
                    detectlocation.click()
                    time.sleep(0.5)
                    confirmbutton = wait.until(EC.presence_of_element_located((By.XPATH , '//button[text() = "Confirm" ]')))
                    confirmbutton.click()
                    time.sleep(0.5)
                onroadpricein = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "On Road Price in " ]'))) 
                currentplace = onroadpricein.find_element(By.XPATH , '//span[text() = "On Road Price in " ]/following-sibling::span')
                currentplacename = currentplace.text
                currentplacename = currentplacename.lstrip()
                currentplacename = currentplacename.rstrip()
                print(currentplacename)
                locationwindow =  wait.until(EC.element_to_be_clickable((By.XPATH , '//div[text() = "City" ]')))
                try:
                    locationwindow.click()
                except:    
                    driver.execute_script("arguments[0].scrollIntoView();", locationwindow)
                    driver.execute_script("arguments[0].click();", locationwindow)
                try:

                    currentplacebox = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "'+currentplacename+'" ]')))
                    currentplaceboxclose = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "'+currentplacename+'" ]/following-sibling::span')))
                    time.sleep(0.5)
                    currentplaceboxclose.click()
                    time.sleep(0.5)            
                except:
                    locationwindow.click()
                    currentplacebox = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "'+currentplacename+'" ]')))
                    currentplaceboxclose = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "'+currentplacename+'" ]/following-sibling::span')))
                    time.sleep(0.5)
                    currentplaceboxclose.click()
                    time.sleep(0.5)            
                liinput = wait.until(EC.element_to_be_clickable((By.XPATH , '//input[@placeholder = "Type your Pincode or City"]')))

                if placeorder != 0:
                    placeval = placeorder
                else :    
                    placeval = 0
                for place in self.Places[placeorder:]: 
                    self.row1 = self.row1 + 1
                    
                    locationwindow =  wait.until(EC.element_to_be_clickable((By.XPATH , '//div[text() = "City" ]'))) 
                    if placeval != 0 and self.reruncheck == 1:
                        try:
                            locationwindow.click()
                        except:    
                            driver.execute_script("arguments[0].scrollIntoView();", locationwindow)
                            driver.execute_script("arguments[0].click();", locationwindow)               
                        try:
                            time.sleep(0.5)
                            try :
                                currentplacebox = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "'+self.Places[placeval-1]+'" ]')))
                                currentplaceboxclose = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "'+self.Places[placeval-1]+'" ]/following-sibling::span')))
                            except:
                                currentplacebox = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "'+currentplacename+'" ]')))
                                currentplaceboxclose = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "'+currentplacename+'" ]/following-sibling::span')))

                            
                            currentplaceboxclose.click()
                            time.sleep(0.5)
                            liinput = wait.until(EC.element_to_be_clickable((By.XPATH , '//input[@placeholder = "Type your Pincode or City"]')))
                        except:
                            try:
                                locationwindow.click()
                            except:
                                pass    
                            try :
                                currentplacebox = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "'+self.Places[placeval-1]+'" ]')))
                                currentplaceboxclose = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "'+self.Places[placeval-1]+'" ]/following-sibling::span')))
                            except:
                                currentplacebox = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "'+currentplacename+'" ]')))
                                currentplaceboxclose = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "'+currentplacename+'" ]/following-sibling::span')))
                            
                            currentplaceboxclose.click()
                            time.sleep(0.5)  
                            liinput = wait.until(EC.element_to_be_clickable((By.XPATH , '//input[@placeholder = "Type your Pincode or City"]')))      
                    liinput.send_keys(place) 
                    time.sleep(0.5)
                    allcities = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "All Cities"]')))
            
                    allcityul = allcities.find_element(By.XPATH , '//span[text() = "All Cities"]/following-sibling::ul')
                    cityli = allcityul.find_element(By.XPATH , './li')
                    cityli.click()
                    time.sleep(0.5)
                    try:
                        popareaul = driver.find_element(By.XPATH , '//span[text() = "Popular Areas"]/following-sibling::ul')
                        popli = popareaul.find_element(By.XPATH , './li')
                        popli.click()
                        time.sleep(0.5)
                    except:
                        pass        
                    try:
                        allareaul = driver.find_element(By.XPATH , '//span[text() = "All Areas"]/following-sibling::ul')
                        allareali = allareaul.find_element(By.XPATH , './li')
                        allareali.click()
                        time.sleep(0.5)
                    except:
                        pass               
                    closepopup = wait.until(EC.presence_of_element_located((By.XPATH , '//span[@aria-label = "Close Popup"]')))#//descentant::svg')))
                    try:
                        closepopup.click()
                    except:
                        pass    
                    print(i.find('a')['title'],'',place)
                    # Variantbutton = wait.until(EC.presence_of_element_located((By.XPATH , '//div[text() = "Variant"]')))
                    Variantbutton = wait.until(EC.presence_of_element_located((By.XPATH , '//p[text() = "Variant"]')))
                    
                    onroadpriceinancestor = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "On Road Price in " ]//ancestor::td'))) 
                    onroadpriceval = onroadpriceinancestor.find_element(By.XPATH , './following-sibling::td')
                    
                    if onroadpriceval.get_attribute('innerHTML') == '₹ N/A':
                        print('This place dont have this bike')
                        self.sheet2.cell(row = self.row1 , column = 1).value = i.find('a')['title']
                        self.sheet2.cell(row = self.row1 , column = 2).value = 'NA'
                        self.sheet2.cell(row = self.row1 , column = 3).value = 'NA'
                        self.sheet2.cell(row = self.row1 , column = 4).value = 'NA'
                        self.sheet2.cell(row = self.row1 , column = 5).value = place                
                        placeval = placeval + 1
                        self.lastbike = i.find('a')['title']
                        self.lastplace = place       
                        self.reruncheck = 1        
                        continue     
                    variantname = Variantbutton.find_element(By.XPATH , './following-sibling::div')
                    startvariant = variantname.get_attribute('innerHTML')
                    exshowroomtext = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "Ex-showroom"]/parent::td')))
                    exshowroomparent = exshowroomtext.find_element(By.XPATH , './ancestor::tbody')
                    for pricecomponents in exshowroomparent.find_elements_by_tag_name('tr'):
                        pricename = pricecomponents.find_element(By.XPATH , './td[1]/span')
                        priceval = pricecomponents.find_element(By.XPATH , './td[2]')
                        self.sheet2.cell(row = self.row1 , column = 1).value = i.find('a')['title']
                        self.sheet2.cell(row = self.row1 , column = 2).value = variantname.get_attribute('innerHTML')
                        self.sheet2.cell(row = self.row1 , column = 3).value = pricename.get_attribute('innerHTML')
                        self.sheet2.cell(row = self.row1 , column = 4).value = priceval.get_attribute('innerHTML')
                        self.sheet2.cell(row = self.row1 , column = 5).value = place 
                        self.row1 = self.row1 + 1  
                                    
                    self.sheet2.cell(row = self.row1-1 , column = 3).value = 'On-road Price' 

                    try:
                        Variantbutton.click()
                    except:    
                        driver.execute_script("arguments[0].scrollIntoView();", Variantbutton)
                        driver.execute_script("arguments[0].click();", Variantbutton)    
                    time.sleep(0.5)
                    
                    for ultag in driver.find_elements_by_tag_name('ul'): 

                        try:
                            liver = ultag.find_element(By.XPATH , './li')
                            verid = liver.get_attribute('verid')
                            if verid != None:
                                break
                        except:
                            continue    
                
                    ulofbikes = wait.until(EC.presence_of_element_located((By.XPATH , '//li[@verid ="'+verid+'"]//parent::ul' )))
                    variantlist = []
                    for variantli in ulofbikes.find_elements_by_tag_name('li'):
                        if variantli.get_attribute('verid') != None:
                            variantlist.append(variantli)
                    if len(variantlist) > 1:
                        
                        for variantli in variantlist:
                            
                            currentvariantname  = variantli.find_element(By.XPATH , './div/div/p')
                            if currentvariantname.get_attribute('innerHTML') == startvariant:
                                variantlist.remove(variantli)
                                break
                        n= 0      
                   
                        for variantli in variantlist:
                            n = n+1
                            ulofbikes = wait.until(EC.presence_of_element_located((By.XPATH , '//li[@verid ="'+verid+'"]//parent::ul' )))
                            variantli.click()
                            onroadpriceinancestor = wait.until(EC.presence_of_element_located((By.XPATH , '//span[text() = "On Road Price in " ]//ancestor::td'))) 
                            onroadpriceval = onroadpriceinancestor.find_element(By.XPATH , './following-sibling::td')                            
                            if onroadpriceval.get_attribute('innerHTML') == '₹ N/A':
                                print('This place dont have this bike')
                                self.sheet2.cell(row = self.row1 , column = 1).value = i.find('a')['title']
                                self.sheet2.cell(row = self.row1 , column = 2).value = 'NA'
                                self.sheet2.cell(row = self.row1 , column = 3).value = 'NA'
                                self.sheet2.cell(row = self.row1 , column = 4).value = 'NA'
                                self.sheet2.cell(row = self.row1 , column = 5).value = place                
                                self.row1 = self.row1 + 1    
                                if n != len(variantlist):
                                    try:
                                        Variantbutton.click()
                                    except:    
                                        driver.execute_script("arguments[0].scrollIntoView();", Variantbutton)
                                        driver.execute_script("arguments[0].click();", Variantbutton) 
                                    time.sleep(0.5)                                  
                                continue                                
                            time.sleep(0.5)
                            for pricecomponents in exshowroomparent.find_elements_by_tag_name('tr'):
                                pricename = pricecomponents.find_element(By.XPATH , './td[1]/span')
                                priceval = pricecomponents.find_element(By.XPATH , './td[2]')
                                self.sheet2.cell(row = self.row1 , column = 1).value = i.find('a')['title']
                                self.sheet2.cell(row = self.row1 , column = 2).value = variantname.get_attribute('innerHTML')
                                self.sheet2.cell(row = self.row1 , column = 3).value = pricename.get_attribute('innerHTML')
                                self.sheet2.cell(row = self.row1 , column = 4).value = priceval.get_attribute('innerHTML')
                                self.sheet2.cell(row = self.row1 , column = 5).value = place 
                                self.row1 = self.row1 + 1        
                            self.sheet2.cell(row = self.row1-1 , column = 3).value = 'On-road Price'    
                            if n != len(variantlist):
                                try:
                                    self.row1 = self.row1 + 1 
                                    Variantbutton.click()
                                except:   
                                    self.row1 = self.row1 + 1  
                                    driver.execute_script("arguments[0].scrollIntoView();", Variantbutton)
                                    driver.execute_script("arguments[0].click();", Variantbutton) 
                                time.sleep(0.5)    
                    else :
                        closepopupvariant = wait.until(EC.presence_of_element_located((By.XPATH , '//span[@aria-label = "Close Popup"]')))  
                        closepopupvariant.click()      

                    print(i.find('a')['title'],' reached')
                    placeval = placeval + 1
                    self.row1 = self.row1 + 1 
                    self.lastbike = i.find('a')['title']
                    self.lastplace = place
                    self.reruncheck = 1
                self.listofbikeback = self.listofbikeback + 1 


    Instance = driverclass()
    driver = Instance.create_driver()
    # Instance.mainfunc(driver,brand)
    try :
        Instance.mainfunc(driver,brand)
    except :
        while True:
            try :
               if len(Instance.listofbikename) > 0 and len(Instance.Places)>0:
                   break
               else :

                   Instance.driver.quit()
                   driver = Instance.create_driver()
                   Instance.mainfunc(driver,brand)   
            except :
                pass

        while Instance.lastbike !=   Instance.listofbikename[-1] or Instance.lastplace != Instance.Places[-1]   :        
            try :
                print(Instance.lastbike,Instance.listofbikename[-1],Instance.lastplace,Instance.Places[-1] ) 
                if Instance.lastplace == Instance.Places[-1]:
                    Instance.lastplace = Instance.Places[0]
                    for j, jname in enumerate(Instance.listofbikename):
                        if jname == Instance.lastbike:
                            Instance.lastbike = Instance.listofbikename[j+1] 
                            break
                print(Instance.lastbike,Instance.listofbikename[-1],Instance.lastplace,Instance.Places[-1] )            
                print('start rerun ',brand)
                Instance.reruncheck = 0
                Instance.row1 = Instance.row1 + 2
                for li,listbike in enumerate(Instance.listofbikename):
                    if listbike == Instance.lastbike:
                        Instance.listofbikeback = li + 1
                        break
                Instance.driver.quit()
                driver = Instance.create_driver()
                Instance.mainfunc(driver,brand)            
            except :
                pass    
    Instance.workbook.save(filename=r'C:\Users\abhishekkd\Documents\Competitor Price'+brand+'3.xlsx')        
    Instance.driver.quit()

    limiter.release()
limiter = BoundedSemaphore(1)
thread = []

for s in Brands:
    
    t = Thread(target=threadtry, args=(s,))
    thread.append(t)
    t.start()
    

for o in thread:
    o.join() 